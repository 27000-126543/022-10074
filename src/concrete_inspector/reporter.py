import os
import csv
from datetime import datetime
from pathlib import Path
from typing import List, Optional
import json

from tabulate import tabulate

from .models import PouringRecord, Severity, Issue
from .validator import get_statistics
from .rules import InspectionRules, ISSUE_TYPE_TO_KEY


SEVERITY_SORT_ORDER = {
    Severity.CRITICAL: 0,
    Severity.HIGH: 1,
    Severity.MEDIUM: 2,
    Severity.LOW: 3,
}


class ReportGenerator:
    def __init__(self, output_dir: Optional[str] = None, rules: Optional[InspectionRules] = None):
        self.output_dir = Path(output_dir) if output_dir else Path.cwd()
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.rules = rules

    def generate_summary_table(self, records: List[PouringRecord], show_all: bool = False) -> str:
        display_records = records if show_all else [r for r in records if r.has_issues]
        if not display_records:
            return "没有符合条件的记录。"

        rows = []
        for r in display_records:
            sev_text = r.highest_severity.value if r.highest_severity else "合格"
            date_str = r.pouring_date.strftime("%Y-%m-%d") if r.pouring_date else "未知"
            issue_brief = "; ".join([i.issue_type.value for i in r.issues[:3]])
            if len(r.issues) > 3:
                issue_brief += f" 等共{len(r.issues)}项"
            consistency_ok = "✅" if not r.consistency_issues else f"⚠️{len(r.consistency_issues)}项"
            rows.append([
                r.record_id[:30],
                r.building or "—",
                date_str,
                r.position or "—",
                r.strength_grade or "—",
                r.supervisor or "—",
                "有" if r.has_supervisor_sign else "无",
                len(r.photos),
                consistency_ok,
                sev_text,
                issue_brief or "—",
            ])

        headers = [
            "记录编号", "楼栋", "浇筑日期", "部位",
            "强度等级", "监理员", "签名", "照片数", "一致性", "严重程度", "问题摘要"
        ]
        return tabulate(rows, headers=headers, tablefmt="grid", showindex=False, stralign="left")

    def generate_issue_detail_table(self, records: List[PouringRecord]) -> str:
        rows = []
        idx = 1
        for r in sorted(records, key=lambda x: (x.highest_severity.order if x.highest_severity else 99, x.record_id)):
            if not r.issues:
                continue
            date_str = r.pouring_date.strftime("%Y-%m-%d") if r.pouring_date else "未知"
            for issue in sorted(r.issues, key=lambda i: SEVERITY_SORT_ORDER.get(i.severity, 99)):
                rows.append([
                    idx,
                    issue.severity.value,
                    r.record_id[:30],
                    r.building or "—",
                    date_str,
                    r.position or "—",
                    r.supervisor or "—",
                    issue.issue_type.value,
                    issue.description,
                    issue.responsible or "—",
                ])
                idx += 1

        if not rows:
            return "没有问题记录。"

        headers = [
            "序号", "严重程度", "记录编号", "楼栋",
            "浇筑日期", "部位", "监理员", "问题类型", "详细描述", "责任人"
        ]
        return tabulate(rows, headers=headers, tablefmt="grid", showindex=False, stralign="left")

    def generate_consistency_table(self, records: List[PouringRecord]) -> str:
        conflicts = [r for r in records if r.consistency_issues]
        if not conflicts:
            return "✅ 所有记录的资料一致性检查通过（楼栋/部位/强度/监理员/日期均无冲突）。"

        rows = []
        idx = 1
        for r in conflicts:
            date_str = r.pouring_date.strftime("%Y-%m-%d") if r.pouring_date else "未知"
            for issue in r.consistency_issues:
                sv = issue.source_values or {}
                rows.append([
                    idx,
                    r.record_id[:30],
                    r.building or "—",
                    date_str,
                    issue.field_name,
                    sv.get("folder") or "—",
                    sv.get("log") or "—",
                    sv.get("manifest") or "—",
                    issue.responsible or "—",
                ])
                idx += 1

        headers = [
            "序号", "记录编号", "楼栋", "浇筑日期", "冲突字段",
            "文件夹名取值", "日志取值", "清单取值", "责任人"
        ]
        return tabulate(rows, headers=headers, tablefmt="grid", showindex=False, stralign="left")

    def generate_action_list(self, records: List[PouringRecord]) -> str:
        issues_only = [r for r in records if r.has_issues]
        if not issues_only:
            return "🎉 所有旁站资料完整，无需整改！"

        issues_only.sort(key=lambda x: (x.highest_severity.order if x.highest_severity else 99, x.record_id))

        sections = {Severity.CRITICAL: [], Severity.HIGH: [], Severity.MEDIUM: [], Severity.LOW: []}

        for r in issues_only:
            hs = r.highest_severity
            if hs and hs in sections:
                sections[hs].append(r)

        lines = []
        lines.append("=" * 80)
        lines.append("混凝土浇筑旁站资料整改清单")
        lines.append(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        if self.rules and self.rules.source_file:
            lines.append(f"规则来源：{self.rules.source_file}")
        lines.append("=" * 80)
        lines.append("")

        severity_titles = {
            Severity.CRITICAL: "🔴 一、严重问题（立即整改，涉及资料有效性）",
            Severity.HIGH: "🟠 二、重要问题（本周内补全）",
            Severity.MEDIUM: "🟡 三、一般问题（3日内补全）",
            Severity.LOW: "🟢 四、轻微问题（尽快完善）",
        }

        total_actions = 0
        for sev in [Severity.CRITICAL, Severity.HIGH, Severity.MEDIUM, Severity.LOW]:
            records_of_sev = sections[sev]
            if not records_of_sev:
                continue
            lines.append(severity_titles[sev])
            lines.append("-" * 80)
            lines.append("")
            for i, r in enumerate(records_of_sev, 1):
                date_str = r.pouring_date.strftime("%Y-%m-%d") if r.pouring_date else "未知日期"
                lines.append(f"  第{i}项  【{r.building or '未知楼栋'}】{r.position or '未知部位'}（{date_str}）")
                lines.append(f"          记录编号：{r.record_id}")
                lines.append(f"          监理员：{r.supervisor or '未填写'}")
                lines.append(f"          待整改内容：")
                for issue in sorted(r.issues, key=lambda x: SEVERITY_SORT_ORDER.get(x.severity, 99)):
                    marker = "●" if issue.severity == sev else "○"
                    lines.append(f"            {marker} [{issue.severity.value}] {issue.description}")
                    if issue.suggestion:
                        lines.append(f"                 → 整改建议：{issue.suggestion}")
                    if issue.responsible:
                        lines.append(f"                 → 责任岗位：{issue.responsible}")
                    total_actions += 1
                lines.append("")

        conflicts = [r for r in issues_only if r.consistency_issues]
        if conflicts:
            lines.append("⚠️  三、资料一致性问题汇总")
            lines.append("-" * 80)
            lines.append("  下列记录在文件夹名、旁站日志、浇筑清单之间存在字段不一致，需核对统一：")
            for r in conflicts:
                for ci in r.consistency_issues:
                    lines.append(f"    - {r.record_id}：{ci.description}")
            lines.append("")

        lines.append("-" * 80)
        lines.append(f"📋 统计：共 {len(issues_only)} 条记录存在问题，需整改事项 {total_actions} 项")
        lines.append("")
        lines.append("责任单位：项目部质量部 / 相关监理组")
        lines.append("整改要求：")
        lines.append("  1. 严重问题：请在24小时内完成补签或补充资料")
        lines.append("  2. 重要问题：请在本周内完成补全")
        lines.append("  3. 一般问题：请在3个工作日内补全")
        lines.append("  4. 所有整改完成后请将补充资料扫描件回复至质量部存档")
        lines.append("=" * 80)
        return "\n".join(lines)

    def generate_statistics_report(self, records: List[PouringRecord]) -> str:
        stats = get_statistics(records)
        lines = []
        lines.append("=" * 60)
        lines.append("混凝土浇筑旁站资料检查统计报告")
        lines.append(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        if self.rules and self.rules.source_file:
            lines.append(f"规则配置：{self.rules.source_file}")
        if self.rules:
            lines.append(f"最少照片：{self.rules.min_photo_count}张  |  "
                         f"时间间隔≤{self.rules.max_photo_gap_hours}h  |  "
                         f"照片时间±{self.rules.photo_early_days}/{self.rules.photo_late_days}天")
        lines.append("=" * 60)
        lines.append("")
        lines.append("【总体情况】")
        lines.append(f"  总记录数：        {stats['总记录数']}")
        lines.append(f"  合格记录：        {stats['合格记录']}")
        lines.append(f"  有问题记录：      {stats['有问题记录']}")
        lines.append(f"  资料合格率：      {stats['合格率']}")
        lines.append(f"  累计问题数：      {stats['总问题数']}")
        if stats.get('一致性冲突数'):
            lines.append(f"  一致性冲突：      {stats['一致性冲突数']} 项")
        lines.append("")
        lines.append("【问题类型分布】")
        issue_dist = stats['问题类型分布']
        if issue_dist:
            max_val = max(issue_dist.values()) if issue_dist else 1
            for issue_name, count in sorted(issue_dist.items(), key=lambda x: -x[1]):
                bar_len = int(count / max_val * 30) if max_val > 0 else 0
                bar = "█" * bar_len
                lines.append(f"  {issue_name:<14} {count:>4}  {bar}")
        else:
            lines.append("  （无问题）")
        lines.append("")
        lines.append("【严重程度分布】")
        sev_dist = stats['严重程度分布']
        for sev_name, count in sev_dist.items():
            lines.append(f"  {sev_name:<6} {count:>4} 项")
        lines.append("")
        lines.append("【范围信息】")
        lines.append(f"  涉及楼栋数：    {stats['涉及楼栋数']}")
        if stats['楼栋列表']:
            lines.append(f"  楼栋：{', '.join(stats['楼栋列表'])}")
        lines.append(f"  涉及监理员数：  {stats['涉及监理员数']}")
        if stats['监理员列表']:
            lines.append(f"  监理员：{', '.join(stats['监理员列表'])}")
        lines.append("")
        lines.append("=" * 60)
        return "\n".join(lines)

    def save_text_report(self, records: List[PouringRecord], filename: str = None) -> str:
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"旁站检查报告_{timestamp}.txt"
        filepath = self.output_dir / filename

        content_parts = []
        content_parts.append(self.generate_statistics_report(records))
        content_parts.append("")
        content_parts.append("📌 问题记录概览：")
        content_parts.append(self.generate_summary_table(records, show_all=False))
        content_parts.append("")
        content_parts.append("🔗 资料一致性核对：")
        content_parts.append(self.generate_consistency_table(records))
        content_parts.append("")
        content_parts.append("📋 问题详细清单：")
        content_parts.append(self.generate_issue_detail_table(records))
        content_parts.append("")
        content_parts.append(self.generate_action_list(records))

        full_content = "\n\n".join(content_parts)
        filepath.write_text(full_content, encoding="utf-8")
        return str(filepath)

    def save_excel_report(self, records: List[PouringRecord], filename: str = None) -> str:
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"旁站检查报告_{timestamp}.xlsx"
        filepath = self.output_dir / filename

        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            print("警告：缺少pandas或openpyxl，无法生成Excel报告。请使用TXT或CSV格式。")
            return ""

        wb = Workbook()
        severity_colors = {
            "严重": "FFC7CE",
            "重要": "FFEB9C",
            "一般": "FFD699",
            "轻微": "C6EFCE",
        }

        ws1 = wb.active
        ws1.title = "统计概览"
        stats = get_statistics(records)
        stats_rows = [
            ["指标", "数值"],
            ["总记录数", stats["总记录数"]],
            ["合格记录", stats["合格记录"]],
            ["有问题记录", stats["有问题记录"]],
            ["合格率", stats["合格率"]],
            ["累计问题数", stats["总问题数"]],
            ["一致性冲突数", stats.get("一致性冲突数", 0)],
            ["涉及楼栋数", stats["涉及楼栋数"]],
            ["涉及监理员数", stats["涉及监理员数"]],
            ["楼栋列表", ", ".join(stats["楼栋列表"])],
            ["监理员列表", ", ".join(stats["监理员列表"])],
        ]
        for row in stats_rows:
            ws1.append(row)
        for cell in ws1[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")

        ws2 = wb.create_sheet("问题记录概览")
        overview_data = [r.to_dict() for r in records]
        if overview_data:
            df = pd.DataFrame(overview_data)
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                ws2.append(row)
            for cell in ws2[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="DDDDDD")

        ws3 = wb.create_sheet("详细问题清单")
        detail_rows = [["序号", "严重程度", "记录编号", "楼栋", "浇筑日期", "部位",
                        "强度等级", "监理员", "问题类型", "详细描述", "整改建议", "责任人"]]
        idx = 1
        for r in sorted(records, key=lambda x: (x.highest_severity.order if x.highest_severity else 99, x.record_id)):
            if not r.issues:
                continue
            date_str = r.pouring_date.strftime("%Y-%m-%d") if r.pouring_date else "未知"
            for issue in sorted(r.issues, key=lambda i: SEVERITY_SORT_ORDER.get(i.severity, 99)):
                detail_rows.append([
                    idx,
                    issue.severity.value,
                    r.record_id,
                    r.building or "—",
                    date_str,
                    r.position or "—",
                    r.strength_grade or "—",
                    r.supervisor or "—",
                    issue.issue_type.value,
                    issue.description,
                    issue.suggestion or "",
                    issue.responsible or "",
                ])
                idx += 1
        for row in detail_rows:
            ws3.append(row)
        for cell in ws3[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
        for row in ws3.iter_rows(min_row=2):
            sev_cell = row[1]
            color = severity_colors.get(str(sev_cell.value))
            if color:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=color)

        ws4 = wb.create_sheet("一致性冲突")
        conflict_rows = [["序号", "记录编号", "楼栋", "浇筑日期", "冲突字段",
                          "文件夹名", "旁站日志", "浇筑清单", "整改建议", "责任人"]]
        ci = 1
        for r in sorted(records, key=lambda x: x.record_id):
            if not r.consistency_issues:
                continue
            date_str = r.pouring_date.strftime("%Y-%m-%d") if r.pouring_date else "未知"
            for issue in r.consistency_issues:
                sv = issue.source_values or {}
                conflict_rows.append([
                    ci, r.record_id, r.building or "—", date_str,
                    issue.field_name,
                    sv.get("folder") or "—", sv.get("log") or "—", sv.get("manifest") or "—",
                    issue.suggestion or "", issue.responsible or "",
                ])
                ci += 1
        for row in conflict_rows:
            ws4.append(row)
        for cell in ws4[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")

        for ws in [ws1, ws2, ws3, ws4]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(str(filepath))
        return str(filepath)

    def save_action_list(self, records: List[PouringRecord], filename: str = None) -> str:
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"整改清单_{timestamp}.txt"
        filepath = self.output_dir / filename
        content = self.generate_action_list(records)
        filepath.write_text(content, encoding="utf-8")
        return str(filepath)

    def save_action_csv(self, records: List[PouringRecord], filename: str = None) -> str:
        """导出适合发项目部的整改清单CSV，含整改建议、责任人和是否整改列"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"整改清单_{timestamp}.csv"
        filepath = self.output_dir / filename

        headers = [
            "序号", "严重程度", "严重程度排序", "记录编号", "楼栋号", "浇筑日期",
            "浇筑部位", "强度等级", "监理员", "问题类型", "问题详情",
            "文件夹原值", "日志原值", "清单原值",
            "整改建议", "责任岗位", "指定责任人（填写）",
            "计划完成日期（填写）", "是否整改（填是/否）", "实际完成日期（填写）", "备注"
        ]

        all_issues = []
        for r in records:
            if not r.issues:
                continue
            date_str = r.pouring_date.strftime("%Y-%m-%d") if r.pouring_date else ""
            for issue in sorted(r.issues, key=lambda i: SEVERITY_SORT_ORDER.get(i.severity, 99)):
                sv = issue.source_values or {}
                all_issues.append((
                    issue, r, date_str,
                    sv.get("folder") or "",
                    sv.get("log") or "",
                    sv.get("manifest") or "",
                ))

        all_issues.sort(key=lambda x: (
            SEVERITY_SORT_ORDER.get(x[0].severity, 99),
            x[1].record_id,
        ))

        try:
            with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for idx, (issue, r, date_str, fv, lv, mv) in enumerate(all_issues, 1):
                    writer.writerow([
                        idx,
                        issue.severity.value,
                        SEVERITY_SORT_ORDER.get(issue.severity, 99),
                        r.record_id,
                        r.building or "",
                        date_str,
                        r.position or "",
                        r.strength_grade or "",
                        r.supervisor or "",
                        issue.issue_type.value,
                        issue.description,
                        fv, lv, mv,
                        issue.suggestion or "",
                        issue.responsible or "",
                        "", "", "否", "", "",
                    ])
        except Exception as e:
            print(f"CSV导出失败：{e}")
            return ""

        return str(filepath)

    def _build_weekly_stats(self, records: List[PouringRecord]) -> dict:
        stats = get_statistics(records)
        total = len(records)
        ok = sum(1 for r in records if not r.has_issues)
        critical = sum(1 for r in records for i in r.issues if i.severity == Severity.CRITICAL)
        high = sum(1 for r in records for i in r.issues if i.severity == Severity.HIGH)
        unmatched = sum(1 for r in records if r.manifest_unmatched)
        consistency = sum(1 for r in records if r.consistency_issues)

        building_stats = {}
        for r in records:
            b = r.building or "未识别楼栋"
            if b not in building_stats:
                building_stats[b] = {"total": 0, "issues": 0, "critical": 0, "issue_count": 0}
            building_stats[b]["total"] += 1
            if r.has_issues:
                building_stats[b]["issues"] += 1
            building_stats[b]["issue_count"] += len(r.issues)
            building_stats[b]["critical"] += len(r.critical_issues)

        supervisor_stats = {}
        for r in records:
            s = r.supervisor or "未填写"
            if s not in supervisor_stats:
                supervisor_stats[s] = {"total": 0, "issues": 0, "issue_count": 0}
            supervisor_stats[s]["total"] += 1
            if r.has_issues:
                supervisor_stats[s]["issues"] += 1
            supervisor_stats[s]["issue_count"] += len(r.issues)

        role_stats = {}
        for r in records:
            for issue in r.issues:
                role = issue.responsible or "未指定"
                role_stats[role] = role_stats.get(role, 0) + 1

        issue_type_stats = {}
        for r in records:
            for issue in r.issues:
                t = issue.issue_type.value
                issue_type_stats[t] = issue_type_stats.get(t, 0) + 1

        return {
            "总览": {
                "总记录数": total,
                "合格记录数": ok,
                "问题记录数": total - ok,
                "资料合格率": f"{(ok/total*100):.1f}%" if total else "0%",
                "累计问题数": stats["总问题数"],
                "严重问题数": critical,
                "重要问题数": high,
                "一致性冲突记录": consistency,
                "未匹配清单数": unmatched,
            },
            "楼栋排行": sorted(building_stats.items(), key=lambda x: -x[1]["issue_count"]),
            "监理员排行": sorted(supervisor_stats.items(), key=lambda x: -x[1]["issue_count"]),
            "问题类型分布": sorted(issue_type_stats.items(), key=lambda x: -x[1]),
            "责任岗位分布": sorted(role_stats.items(), key=lambda x: -x[1]),
        }

    def generate_weekly_summary(self, records: List[PouringRecord]) -> str:
        w = self._build_weekly_stats(records)
        lines = []
        lines.append("=" * 70)
        lines.append("混凝土浇筑旁站检查  每周抽查汇总报告")
        lines.append(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        if self.rules and self.rules.source_file:
            lines.append(f"规则来源：{self.rules.source_file}")
        lines.append("=" * 70)
        lines.append("")

        lines.append("【一、本周总体情况】")
        lines.append("-" * 70)
        ov = w["总览"]
        for k, v in ov.items():
            lines.append(f"  {k:<16} {v}")
        lines.append("")

        lines.append("【二、楼栋问题排行（按问题总数降序）】")
        lines.append("-" * 70)
        lines.append(f"  {'排名':<4} {'楼栋':<12} {'记录数':<8} {'问题记录':<8} {'问题总数':<8} {'严重问题':<8}")
        for idx, (b, s) in enumerate(w["楼栋排行"], 1):
            lines.append(f"  {idx:<4} {b:<12} {s['total']:<8} {s['issues']:<8} {s['issue_count']:<8} {s['critical']:<8}")
        lines.append("")

        lines.append("【三、监理员问题排行】")
        lines.append("-" * 70)
        lines.append(f"  {'排名':<4} {'监理员':<12} {'记录数':<8} {'问题记录':<8} {'问题总数':<8}")
        for idx, (s, st) in enumerate(w["监理员排行"], 1):
            lines.append(f"  {idx:<4} {s:<12} {st['total']:<8} {st['issues']:<8} {st['issue_count']:<8}")
        lines.append("")

        lines.append("【四、问题类型分布】")
        lines.append("-" * 70)
        max_val = max([v for _, v in w["问题类型分布"]] + [1])
        for t, c in w["问题类型分布"]:
            bar = "█" * int(c / max_val * 30)
            lines.append(f"  {t:<18} {c:>4}  {bar}")
        lines.append("")

        lines.append("【五、责任岗位分布】")
        lines.append("-" * 70)
        max_r = max([v for _, v in w["责任岗位分布"]] + [1])
        for role, c in w["责任岗位分布"]:
            bar = "█" * int(c / max_r * 30)
            lines.append(f"  {role:<18} {c:>4}  {bar}")
        lines.append("")

        lines.append("【六、重点关注问题】")
        lines.append("-" * 70)
        high_priority = []
        for r in records:
            for issue in r.issues:
                if issue.severity in (Severity.CRITICAL, Severity.HIGH):
                    high_priority.append((issue, r))
        high_priority.sort(key=lambda x: (x[0].severity.order, x[1].record_id))
        if high_priority:
            for idx, (issue, r) in enumerate(high_priority[:15], 1):
                date_str = r.pouring_date.strftime("%Y-%m-%d") if r.pouring_date else "未知"
                lines.append(f"  {idx:>2}. [{issue.severity.value}] {r.building or '-'} {r.position or '-'}（{date_str}）")
                lines.append(f"      · {issue.description}")
                lines.append(f"      · 责任岗位：{issue.responsible or '-'}")
            if len(high_priority) > 15:
                lines.append(f"  ... 共 {len(high_priority)} 项，仅显示前15项")
        else:
            lines.append("  ✅ 暂无严重/重要问题")
        lines.append("")

        unmatched_records = [r for r in records if r.manifest_unmatched]
        if unmatched_records:
            lines.append("【七、未匹配清单记录】")
            lines.append("-" * 70)
            lines.append("  以下清单条目未找到对应文件夹，请核对：")
            for r in unmatched_records:
                date_str = r.pouring_date.strftime("%Y-%m-%d") if r.pouring_date else "未知"
                lines.append(f"  · {r.building or '未知楼栋'} {r.position or '未知部位'}（{date_str}） - {r.strength_grade or '未知强度'}")
            lines.append("")

        lines.append("=" * 70)
        return "\n".join(lines)

    def save_weekly_csv(self, records: List[PouringRecord], filename: str = None) -> str:
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"周报汇总_{timestamp}.csv"
        filepath = self.output_dir / filename

        w = self._build_weekly_stats(records)

        try:
            with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)

                writer.writerow(["混凝土浇筑旁站每周抽查汇总报告"])
                writer.writerow([f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                if self.rules and self.rules.source_file:
                    writer.writerow([f"规则来源：{self.rules.source_file}"])
                writer.writerow([])

                writer.writerow(["一、总体情况"])
                writer.writerow(["指标", "数值"])
                for k, v in w["总览"].items():
                    writer.writerow([k, v])
                writer.writerow([])

                writer.writerow(["二、楼栋问题排行"])
                writer.writerow(["排名", "楼栋", "记录数", "问题记录数", "问题总数", "严重问题数", "合格率"])
                for idx, (b, s) in enumerate(w["楼栋排行"], 1):
                    rate = f"{(1-s['issues']/s['total'])*100:.1f}%" if s['total'] else "0%"
                    writer.writerow([idx, b, s["total"], s["issues"], s["issue_count"], s["critical"], rate])
                writer.writerow([])

                writer.writerow(["三、监理员问题排行"])
                writer.writerow(["排名", "监理员", "记录数", "问题记录数", "问题总数"])
                for idx, (s, st) in enumerate(w["监理员排行"], 1):
                    writer.writerow([idx, s, st["total"], st["issues"], st["issue_count"]])
                writer.writerow([])

                writer.writerow(["四、问题类型分布"])
                writer.writerow(["问题类型", "数量", "占比"])
                total_issues = sum(v for _, v in w["问题类型分布"])
                for t, c in w["问题类型分布"]:
                    pct = f"{c/total_issues*100:.1f}%" if total_issues else "0%"
                    writer.writerow([t, c, pct])
                writer.writerow([])

                writer.writerow(["五、责任岗位分布"])
                writer.writerow(["责任岗位", "问题数", "占比"])
                total_role = sum(v for _, v in w["责任岗位分布"])
                for role, c in w["责任岗位分布"]:
                    pct = f"{c/total_role*100:.1f}%" if total_role else "0%"
                    writer.writerow([role, c, pct])
                writer.writerow([])

                writer.writerow(["六、未匹配清单记录"])
                writer.writerow(["序号", "楼栋", "部位", "强度等级", "浇筑日期", "监理员"])
                unmatched = [r for r in records if r.manifest_unmatched]
                for idx, r in enumerate(unmatched, 1):
                    date_str = r.pouring_date.strftime("%Y-%m-%d") if r.pouring_date else ""
                    writer.writerow([idx, r.building or "", r.position or "", r.strength_grade or "", date_str, r.supervisor or ""])

        except Exception as e:
            print(f"周报CSV导出失败：{e}")
            return ""
        return str(filepath)

    def save_weekly_excel(self, records: List[PouringRecord], filename: str = None) -> str:
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"周报汇总_{timestamp}.xlsx"
        filepath = self.output_dir / filename

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.chart import BarChart, Reference
        except ImportError:
            print("警告：缺少openpyxl，无法生成Excel周报。")
            return ""

        w = self._build_weekly_stats(records)
        wb = Workbook()

        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")
        sub_fill = PatternFill("solid", fgColor="D9E2F3")
        sub_font = Font(bold=True)

        ws1 = wb.active
        ws1.title = "周报总览"

        row = 1
        ws1.cell(row=row, column=1, value="混凝土浇筑旁站每周抽查汇总报告").font = Font(bold=True, size=14)
        row += 1
        ws1.cell(row=row, column=1, value=f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        row += 1
        if self.rules and self.rules.source_file:
            ws1.cell(row=row, column=1, value=f"规则来源：{self.rules.source_file}")
            row += 1
        row += 1

        ws1.cell(row=row, column=1, value="一、总体情况").fill = sub_fill
        ws1.cell(row=row, column=1).font = sub_font
        row += 1
        for k, v in w["总览"].items():
            ws1.cell(row=row, column=1, value=k)
            ws1.cell(row=row, column=2, value=v)
            row += 1
        row += 1

        ws1.cell(row=row, column=1, value="二、楼栋问题排行").fill = sub_fill
        ws1.cell(row=row, column=1).font = sub_font
        row += 1
        headers = ["排名", "楼栋", "记录数", "问题记录数", "问题总数", "严重问题数", "合格率"]
        for col, h in enumerate(headers, 1):
            c = ws1.cell(row=row, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
        row += 1
        for idx, (b, s) in enumerate(w["楼栋排行"], 1):
            rate = f"{(1-s['issues']/s['total'])*100:.1f}%" if s['total'] else "0%"
            ws1.cell(row=row, column=1, value=idx)
            ws1.cell(row=row, column=2, value=b)
            ws1.cell(row=row, column=3, value=s["total"])
            ws1.cell(row=row, column=4, value=s["issues"])
            ws1.cell(row=row, column=5, value=s["issue_count"])
            ws1.cell(row=row, column=6, value=s["critical"])
            ws1.cell(row=row, column=7, value=rate)
            row += 1
        row += 2

        ws1.cell(row=row, column=1, value="三、监理员问题排行").fill = sub_fill
        ws1.cell(row=row, column=1).font = sub_font
        row += 1
        headers2 = ["排名", "监理员", "记录数", "问题记录数", "问题总数"]
        for col, h in enumerate(headers2, 1):
            c = ws1.cell(row=row, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
        row += 1
        for idx, (s, st) in enumerate(w["监理员排行"], 1):
            ws1.cell(row=row, column=1, value=idx)
            ws1.cell(row=row, column=2, value=s)
            ws1.cell(row=row, column=3, value=st["total"])
            ws1.cell(row=row, column=4, value=st["issues"])
            ws1.cell(row=row, column=5, value=st["issue_count"])
            row += 1

        for col in range(1, 8):
            ws1.column_dimensions[chr(64+col)].width = 16

        ws2 = wb.create_sheet("问题类型分布")
        ws2.cell(row=1, column=1, value="问题类型").fill = header_fill
        ws2.cell(row=1, column=2, value="数量").fill = header_fill
        ws2.cell(row=1, column=1).font = header_font
        ws2.cell(row=1, column=2).font = header_font
        for idx, (t, c) in enumerate(w["问题类型分布"], 2):
            ws2.cell(row=idx, column=1, value=t)
            ws2.cell(row=idx, column=2, value=c)
        ws2.column_dimensions['A'].width = 22
        ws2.column_dimensions['B'].width = 12

        try:
            chart = BarChart()
            chart.type = "bar"
            chart.style = 10
            chart.title = "问题类型分布"
            data = Reference(ws2, min_col=2, min_row=1, max_row=1+len(w["问题类型分布"]))
            cats = Reference(ws2, min_col=1, min_row=2, max_row=1+len(w["问题类型分布"]))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 12
            chart.width = 20
            ws2.add_chart(chart, "D2")
        except Exception:
            pass

        ws3 = wb.create_sheet("责任岗位分布")
        ws3.cell(row=1, column=1, value="责任岗位").fill = header_fill
        ws3.cell(row=1, column=2, value="问题数").fill = header_fill
        ws3.cell(row=1, column=1).font = header_font
        ws3.cell(row=1, column=2).font = header_font
        for idx, (role, c) in enumerate(w["责任岗位分布"], 2):
            ws3.cell(row=idx, column=1, value=role)
            ws3.cell(row=idx, column=2, value=c)
        ws3.column_dimensions['A'].width = 20
        ws3.column_dimensions['B'].width = 12

        try:
            chart2 = BarChart()
            chart2.type = "bar"
            chart2.style = 11
            chart2.title = "责任岗位问题分布"
            data2 = Reference(ws3, min_col=2, min_row=1, max_row=1+len(w["责任岗位分布"]))
            cats2 = Reference(ws3, min_col=1, min_row=2, max_row=1+len(w["责任岗位分布"]))
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(cats2)
            chart2.height = 10
            chart2.width = 18
            ws3.add_chart(chart2, "D2")
        except Exception:
            pass

        ws4 = wb.create_sheet("未匹配清单")
        headers3 = ["序号", "楼栋", "部位", "强度等级", "浇筑日期", "监理员"]
        for col, h in enumerate(headers3, 1):
            c = ws4.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
        unmatched = [r for r in records if r.manifest_unmatched]
        for idx, r in enumerate(unmatched, 2):
            date_str = r.pouring_date.strftime("%Y-%m-%d") if r.pouring_date else ""
            ws4.cell(row=idx, column=1, value=idx-1)
            ws4.cell(row=idx, column=2, value=r.building or "")
            ws4.cell(row=idx, column=3, value=r.position or "")
            ws4.cell(row=idx, column=4, value=r.strength_grade or "")
            ws4.cell(row=idx, column=5, value=date_str)
            ws4.cell(row=idx, column=6, value=r.supervisor or "")
        for col in range(1, 7):
            ws4.column_dimensions[chr(64+col)].width = 16

        wb.save(str(filepath))
        return str(filepath)
